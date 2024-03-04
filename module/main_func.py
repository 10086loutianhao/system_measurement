#!/usr/bin/env python
# coding: utf-8

# In[7]:


# set Table menu and set start_column_index and start test , depend on import serial & subprocess & time & pyautogui & csv, 
# import pyscreenshot as ImageGrab , import numpy as np , from IPython.display import clear_output
from menuconfig import set_menucofig_command 
from download import change_time_value_and_download
from serial_contorl import serial_rst_chip_to_boot
from serial_contorl import serial_process_data
from desk_image_process import wait_export_csv
from level_measurement import process_csv_and_write 
from level_measurement import find_max_value_in_csv_sublist
from IPython.display import clear_output
from openpyxl.styles import Alignment
import DSview_ctrl
import pyautogui
import time
import numpy as np

def main_func(DFS_min_list, DFS_max_list, RTC_source_list, times_list, PD_flash_flag_list, csvname, wb, IDF_PATH, EXAMPLE_PATH, \
                 csv_sublist, serial_sublist, dev_sublist, optional_items_list,  \
                 input_pos, target_port, offset, dict_list, baud = 115200):
    
    #############   create tavle menu and return the next start row    ############
    ret_row = set_table_menu(DFS_min_list, DFS_max_list, RTC_source_list, times_list, PD_flash_flag_list, optional_items_list, wb)
        
    counter_1 = 0
    counter_2 = 0
    counter_3 = 0
    mux_1 = len(RTC_source_list) * len(PD_flash_flag_list) * len(optional_items_list)
    mux_2 = len(PD_flash_flag_list) * len(optional_items_list)
    mux_3 = len(optional_items_list)
    #offset = 631
    start_column_index_1 = 0
    start_column_index_2 = 0
    start_column_index_3 = 0
    
    for DFS_min in DFS_min_list:
        dfs = 'dfs '
        dfs = dfs + str(DFS_min) + ' ~ '
        for DFS_max in DFS_max_list:
            dfs_scope = dfs + str(DFS_max)
            ws = wb[dfs_scope]
            for times in times_list:
                start_column_index_1 = 1 + mux_1 * counter_1
                counter_1 += 1 
                counter_1 = counter_1 % len(times_list)
                for RTC_source in RTC_source_list:
                    start_column_index_2 = start_column_index_1 + mux_2 * counter_2
                    counter_2 += 1
                    counter_2 = counter_2 % len(RTC_source_list)
                    for PD_flash_flag in PD_flash_flag_list:
                        start_column_index_3 = start_column_index_2 + mux_3 * counter_3
                        print("DFS_min = {}, DFS_max = {}, times = {}, RTC_source = {}, PD_flash_flag = {}".format(DFS_min, DFS_max, times, RTC_source, PD_flash_flag))
                        #######   call set_menucofig_command    #######
                        set_menucofig_command(DFS_min, DFS_max, PD_flash_flag, RTC_source, IDF_PATH)
                        
                        #######   call change_time_value_and_download then enter download mode   #######   
                        change_time_value_and_download(times, IDF_PATH, EXAMPLE_PATH, target_port, baud)
                        
                        #############   control mouse to use DSview , Wait for trigger    ############
                        DSview_ctrl.click_start(input_pos)
                        
                        ###########   mannually select serial to reset chip and record data     ############
                        serial_rst_chip_to_boot(target_port, baud)

                        #############    process serial data and record it in ws, serial send data sequence should be equivalent to serial_sublist  ##############
                        start_row = ret_row
                        start_col = start_column_index_3
                        ret_serial_dict = search_sublist_pos_in_parentlist(serial_sublist, optional_items_list)
                        end_row = serial_process_data(target_port, ws, start_row, start_col, ret_serial_dict, baud)
                        
                        #############   start to export csv, sampling rate should be 1 M    ##############
                        DSview_ctrl.click_stop(input_pos)
                        DSview_ctrl.click_file(input_pos)
                        DSview_ctrl.click_change(input_pos)
                        time.sleep(0.5)
                        pyautogui.write('1.csv', interval=0.25)
                        time.sleep(0.5)
                        pyautogui.hotkey("enter")
                        time.sleep(0.5)
                        pyautogui.hotkey("enter")
                        DSview_ctrl.click_ok_and_leave(input_pos)
                        
                        #######   Grab the part of the screen to make sure sucessfully export csv   #######
                        wait_export_csv(input_pos)
                        
                        #############    process csv  and record it in ws   ##############
                        ret_csv_dict = search_sublist_pos_in_parentlist(csv_sublist, optional_items_list)
                        print(ret_csv_dict)
                        row_need_skip = process_csv_and_write(csvname, start_row, start_col, ws, ret_csv_dict)
                        print(row_need_skip)
                        tinydict = find_max_value_in_csv_sublist(start_row, end_row, start_col, ws, ret_csv_dict, row_need_skip)
                        
                        #############    calc dev  and record it in ws   ##############
                        ret_dev_dict = search_sublist_pos_in_parentlist(dev_sublist, optional_items_list)
                        tinydict.update(calculate_dev_and_result(start_row, end_row, start_col, offset, ws, ret_dev_dict, row_need_skip))
                                
                        ###################    save results for charts   #################
                        tinydict.update({ 'dfs_min' : DFS_min, 'dfs_max' : DFS_max, 'RTC_source' : RTC_source,\
                                   'times' : times, 'PD_flash_flag' : PD_flash_flag})
                        dict_list.append(tinydict)

                        ##########          end of circle        ############
                        counter_3 += 1
                        counter_3 = counter_3 % len(PD_flash_flag_list)
                        clear_output(wait=True)
    
    


# In[8]:


def column_to_name(colnum):
    if type(colnum) is not int:
        return colnum
    str = ''
    while(not(colnum//26 == 0 and colnum % 26 == 0)):
        temp = 25
        if(colnum % 26 == 0):
            str += chr(temp+65)
        else:
            str += chr(colnum % 26 - 1 + 65)
        colnum //= 26
        #print(str)
    #倒序输出拼写的字符串
    return str[::-1]

def set_table_menu(DFS_min_list, DFS_max_list, RTC_source_list, times_list, PD_flash_flag_list, optional_items_list, wb):
    counter_1 = 0
    counter_2 = 0
    counter_3 = 0
    mux_1 = len(RTC_source_list) * len(PD_flash_flag_list) * len(optional_items_list)
    mux_2 = len(PD_flash_flag_list) * len(optional_items_list)
    mux_3 = len(optional_items_list)
    for DFS_min in DFS_min_list:
        dfs = 'dfs '
        dfs = dfs + str(DFS_min) + ' ~ '
        for DFS_max in DFS_max_list:
            dfs_scope = dfs + str(DFS_max)
            ws = wb.create_sheet(dfs_scope)
            for times in times_list:
                time_str = 'times = 1s * '
                time_str = time_str + str(times)
                ws.merge_cells(start_row = 1, start_column = 1 + mux_1 * counter_1, end_row = 1, end_column = mux_1 * (counter_1 + 1) )
                ws.cell(1, 1 + mux_1 * counter_1, time_str).alignment = Alignment(horizontal='center', vertical='center')
                start_column_index_1 = 1 + mux_1 * counter_1
                counter_1 += 1 
                counter_1 = counter_1 % len(times_list)
                for RTC_source in RTC_source_list:
                    rtc = 'RTC source = '
                    rtc = rtc + RTC_source
                    ws.merge_cells(start_row = 2, start_column = start_column_index_1 + mux_2 * counter_2, end_row = 2, end_column = start_column_index_1 + mux_2 * ( counter_2 + 1 ) - 1 )
                    ws.cell(2, start_column_index_1 + mux_2 * counter_2, rtc).alignment = Alignment(horizontal='center', vertical='center')
                    start_column_index_2 = start_column_index_1 + mux_2 * counter_2
                    counter_2 += 1
                    counter_2 = counter_2 % len(RTC_source_list)
                    for PD_flash_flag in PD_flash_flag_list:
                        pd = ''
                        pd = pd + PD_flash_flag
                        ws.merge_cells(start_row = 3, start_column = start_column_index_2 + mux_3 * counter_3, end_row = 3, end_column = start_column_index_2 + mux_3 * ( counter_3 + 1 ) - 1 )
                        ws.cell(3, start_column_index_2 + mux_3 * counter_3, pd).alignment = Alignment(horizontal='center', vertical='center')
                        option_item_index = 0
                        for option_item in optional_items_list: 
                            ws.cell(4, start_column_index_2 + mux_3 * counter_3 + option_item_index, option_item).alignment = Alignment(horizontal='center', vertical='center')
                            ws.column_dimensions[column_to_name(start_column_index_2 + mux_3 * counter_3 + option_item_index)].width = len(option_item)
                            option_item_index += 1
                        counter_3 += 1
                        counter_3 = counter_3 % len(PD_flash_flag_list)
    #返回待填写数据的行
    return 5    

def search_sublist_pos_in_parentlist(sublist, parentlist):
    ret_dict = {}
    for sublist_item in range(len(sublist)):
        for parentlist_item in range(len(parentlist)):
            if sublist[sublist_item] == parentlist[parentlist_item]:
                ret_dict[str(sublist[sublist_item])] = parentlist_item
    return ret_dict

def calculate_dev_and_result(start_row, end_row, start_col, offset, ws, ret_dev_dict, row_need_skip):
    
    # 获取字典的键列表
    dev_keys = list(ret_dev_dict.keys())
    
    dev_num = int(len(ret_dev_dict) / 3)
    
    dev_maximum = [0] * dev_num
    dev_minimum = [0] * dev_num
    
    for row_num in range(start_row, end_row):
        if row_num in row_need_skip:
            if row_num == start_row:
                start_row += 1
            continue
            
        for item in range(dev_num):
            dev_left = ws.cell(row_num, start_col + ret_dev_dict[dev_keys[item * 3 + 0]]).value
            dev_right = ws.cell(row_num, start_col + ret_dev_dict[dev_keys[item * 3 + 1]]).value
#             print("row_num = ", end="")
#             print(row_num)
#             print("start_col + ret_dev_dict[dev_keys[item * 3 + 0]] = ", end="")
#             print(start_col + ret_dev_dict[dev_keys[item * 3 + 0]])
#             print("dev_left = ", end="")
#             print(dev_left)
#             print("start_col + ret_dev_dict[dev_keys[item * 3 + 1]] = ", end="")
#             print(start_col + ret_dev_dict[dev_keys[item * 3 + 1]])
#             print("dev_right = ", end="")
#             print(dev_right)
            dev = int(dev_left) - int(dev_right)
            ws.cell(row_num, start_col + ret_dev_dict[dev_keys[item * 3 + 2]], dev)
            
            if row_num == start_row:
                dev_maximum[item] = dev
                dev_minimum[item] = dev
            if row_num > start_row:
                if dev_maximum[item] < dev:
                    dev_maximum[item] = dev
                if dev_minimum[item] > dev:
                    dev_minimum[item] = dev
                    
    dev_maximum = [dev_maximum[i] + offset[i] for i in range(len(dev_maximum))]
    dev_minimum = [dev_minimum[i] + offset[i] for i in range(len(dev_minimum))]
    print("dev data maximum and minimum have been recorded in dict and will be utilized for cartography") 
    return {'dev_maximum': dev_maximum, 'dev_minimum': dev_minimum}


# In[ ]:




